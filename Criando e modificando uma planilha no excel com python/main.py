import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

# Criando a página

book.create_sheet('Physics')

# Selecionando a página

Physics_page = book['Physics']
Physics_page.append(['Nome', 'Ano de nascimento', 'Contribuição '])
Physics_page.append(['Newton', '1643', 'Cálculo integral '])
Physics_page.append(['Heinsenberg', '1901', 'Princípio da incerteza '])
Physics_page.append(['Max Plack', '1858', 'Mecância quântica'])

#Salvando a planilha
book.save('Planilha dos físicos.xlsx')

